"""Tests for tools/token_usage_stats.py (батч мелочей после калибровки
№6, D-0081, п.1). Паттерн изоляции -- как в tools/test_usage_report.py:
временная sqlite БД (создана прямым INSERT в cc_usage через
usage_report.SCHEMA, БЕЗ транскриптов), GATEWAY_DB_PATH указывает на
неё (мотив -- инцидент 2026-08-04 16:52, см. докстринг модуля
token_usage_stats.py).

transcript_glob монкипатчена на пустой временный каталог -- import_
transcripts() не сканирует реальный ~/.claude/projects разработчика.

ПЕРЕСДАЧА (блокер 1 критика, живая проба 2026-08-05): ЭТОТ файл БОЛЬШЕ
НЕ патчит tools/token_usage_stats.XLSX_PATH (никакой autouse-фикстуры
на модульный глобал) -- ИМЕННО этот приём (патч мутируемого XLSX_PATH
ради собственной изоляции) СНИМАЛ guard с РЕАЛЬНОГО боевого файла для
любого вызова с явным --output, указывающим буквально на него (см.
докстринг _default_xlsx_path() в token_usage_stats.py за полный разбор
находки критика). Правило теперь простое: КАЖДЫЙ вызов tus.main() в
этом файле несёт ЯВНЫЙ `--output <tmp path>`; поведение guard'а на
РЕАЛЬНОМ дефолтном пути проверяется ТОЛЬКО прямым вызовом чистой
функции `_is_safe_to_write()` (вердикт, не запись -- см. секцию
FAIL-CLOSED GUARD ниже) -- ни один тест этого файла не пишет и не
читает logs/token_usage.xlsx репозитория ни в каком виде.

Run from tools/: python -m pytest test_token_usage_stats.py -q
"""

import sqlite3
from datetime import datetime, timezone
from pathlib import Path

import pytest

import token_usage_stats as tus
import usage_report

FAMILY_SONNET = "claude-sonnet-5"


def _make_db(tmp_path) -> Path:
    db_file = tmp_path / "requests.db"
    conn = sqlite3.connect(db_file)
    conn.execute(usage_report.SCHEMA)
    conn.commit()
    conn.close()
    return db_file


def _insert_row(db_file, ts, model=FAMILY_SONNET, project="projA",
                 session_id="s-1", dedupe_suffix=None,
                 input_tokens=100, output_tokens=50):
    cost, _ = usage_report.accounted_cost(model, input_tokens, output_tokens, 0, 0)
    dedupe = f"{session_id}:{dedupe_suffix or ts}"
    conn = sqlite3.connect(db_file)
    conn.execute(
        """
        INSERT INTO cc_usage
            (ts, project, session_id, turn_index, model, input_tokens,
             output_tokens, cache_creation_tokens, cache_read_tokens,
             accounted_cost_usd, traffic_kind, is_sidechain, dedupe_key)
        VALUES (?, ?, ?, 0, ?, ?, ?, 0, 0, ?, 'real', 0, ?)
        """,
        (ts, project, session_id, model, input_tokens, output_tokens, cost, dedupe),
    )
    conn.commit()
    conn.close()


@pytest.fixture()
def isolated_db(tmp_path, monkeypatch):
    """Изолированная БД (см. докстринг модуля) + пустой transcript_glob
    -- нулевой транскрипт-скан реального ~/.claude/projects."""
    db_file = _make_db(tmp_path)
    monkeypatch.setenv("GATEWAY_DB_PATH", str(db_file))
    empty_dir = tmp_path / "no_transcripts"
    empty_dir.mkdir()
    monkeypatch.setattr(
        tus, "transcript_glob",
        lambda base_dir=None: [str(empty_dir / "*" / "*.jsonl")],
    )
    return db_file


# ---------------------------------------------------------------------
# Оконное бакетирование по границе 14-го (правило 6а -- обе стороны).
# ---------------------------------------------------------------------


def test_window_of_exact_local_midnight_14th_is_new_window():
    # "строка ровно на границе (14-е, 00:00:00 локального времени)
    # относится к НОВОМУ окну" -- буквальный граничный тест.
    assert tus.window_of(datetime(2026, 6, 14, 0, 0, 0)) == (2026, 6)


def test_window_of_one_second_before_midnight_14th_is_old_window():
    assert tus.window_of(datetime(2026, 6, 13, 23, 59, 59)) == (2026, 5)


def test_utc_stored_ts_buckets_by_local_time_not_utc_date(isolated_db):
    # cc_usage.ts хранится в UTC ('Z'); окно бакетируется по ЛОКАЛЬНОМУ
    # времени. Noon UTC ("T12:00:00Z") конвертируется в местное время
    # ЛЮБОЙ реалистичной таймзоны (-12..+14ч) без пересечения границы
    # окна в ОБРАТНУЮ сторону -- тест устойчив к TZ машины прогона (не
    # хардкодит конкретное смещение).
    _insert_row(isolated_db, "2026-06-12T12:00:00Z", session_id="s-old")  # < 14 везде
    _insert_row(isolated_db, "2026-06-14T12:00:00Z", session_id="s-new")  # >= 14 везде

    conn = sqlite3.connect(isolated_db)
    try:
        by_family, _by_project, _sessions = tus.aggregate(conn)
    finally:
        conn.close()

    windows = set(by_family.keys())
    assert (2026, 5) in windows  # старое окно поймало "12-е" строку
    assert (2026, 6) in windows  # новое окно поймало "14-е" строку


# ---------------------------------------------------------------------
# Идемпотентность append -- второй прогон не дублирует период НИ на
# usage, НИ на projects.
# ---------------------------------------------------------------------


def test_append_is_idempotent_on_both_sheets(isolated_db, tmp_path):
    # Завершённый период -- ts заведомо в прошлом (полностью закрытое
    # окно 14.01-14.02.2020).
    _insert_row(isolated_db, "2020-01-20T12:00:00Z", session_id="s-1")
    out = tmp_path / "out.xlsx"

    code1 = tus.main(["--output", str(out)])
    assert code1 == 0
    assert out.exists()

    from openpyxl import load_workbook
    wb1 = load_workbook(out)
    usage_rows_1 = wb1[tus.USAGE_SHEET].max_row
    projects_rows_1 = wb1[tus.PROJECTS_SHEET].max_row

    code2 = tus.main(["--output", str(out)])
    assert code2 == 0

    wb2 = load_workbook(out)
    usage_rows_2 = wb2[tus.USAGE_SHEET].max_row
    projects_rows_2 = wb2[tus.PROJECTS_SHEET].max_row

    assert usage_rows_2 == usage_rows_1
    assert projects_rows_2 == projects_rows_1
    # период не задублирован -- число СТРОК с этим label (family-строки
    # + ИТОГО, wb1 уже даёт эталонное N для одного вхождения периода)
    # не меняется между прогонами.
    label = tus.window_label(2020, 1)
    label_rows_1 = [
        row[0].value for row in wb1[tus.USAGE_SHEET].iter_rows(min_row=1)
        if row[0].value == label
    ]
    label_rows_2 = [
        row[0].value for row in wb2[tus.USAGE_SHEET].iter_rows(min_row=1)
        if row[0].value == label
    ]
    assert len(label_rows_1) >= 1
    assert len(label_rows_2) == len(label_rows_1)
    proj_label_rows_1 = [
        row[0].value for row in wb1[tus.PROJECTS_SHEET].iter_rows(min_row=1)
        if row[0].value == label
    ]
    proj_label_rows_2 = [
        row[0].value for row in wb2[tus.PROJECTS_SHEET].iter_rows(min_row=1)
        if row[0].value == label
    ]
    assert len(proj_label_rows_1) >= 1
    assert len(proj_label_rows_2) == len(proj_label_rows_1)


# ---------------------------------------------------------------------
# Незавершённое (текущее) окно НЕ дописывается.
# ---------------------------------------------------------------------


def test_current_incomplete_window_not_appended(isolated_db, tmp_path):
    now_utc_ts = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    _insert_row(isolated_db, now_utc_ts, session_id="s-now")
    out = tmp_path / "out.xlsx"
    code = tus.main(["--output", str(out)])
    assert code == 0
    # ноль завершённых периодов -- файл НЕ создаётся.
    assert not out.exists()


def test_completed_and_incomplete_windows_together_only_completed_appended(isolated_db, tmp_path):
    _insert_row(isolated_db, "2020-01-20T12:00:00Z", session_id="s-old", dedupe_suffix="a")
    now_utc_ts = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    _insert_row(isolated_db, now_utc_ts, session_id="s-now", dedupe_suffix="b")
    out = tmp_path / "out.xlsx"
    code = tus.main(["--output", str(out)])
    assert code == 0
    assert out.exists()

    from openpyxl import load_workbook
    wb = load_workbook(out)
    periods = {row[0].value for row in wb[tus.USAGE_SHEET].iter_rows(min_row=1) if row[0].value}
    assert tus.window_label(2020, 1) in periods
    # текущее (незавершённое) окно НЕ появилось на листе
    wkey_now = tus.window_of(datetime.now())
    assert tus.window_label(*wkey_now) not in periods


# ---------------------------------------------------------------------
# Пустая БД / ноль завершённых периодов.
# ---------------------------------------------------------------------


def test_empty_db_no_file_created_and_message_printed(isolated_db, tmp_path, capsys):
    out = tmp_path / "out.xlsx"
    code = tus.main(["--output", str(out)])
    assert code == 0
    assert not out.exists()
    captured = capsys.readouterr()
    assert "no new completed periods to append" in captured.out


# ---------------------------------------------------------------------
# FAIL-CLOSED GUARD (п.2 спеки батча). Пересдача (блокер 1 критика):
# поведение guard'а на РЕАЛЬНОМ дефолтном пути доказывается ТОЛЬКО
# вердиктом чистой функции _is_safe_to_write() -- ни один тест не
# вызывает main() без --output и не создаёт/не открывает файл на
# реальном дефолтном пути (см. докстринг модуля выше, "ПЕРЕСДАЧА").
# ---------------------------------------------------------------------


def test_is_safe_to_write_blocks_real_default_path_when_db_swapped(monkeypatch, tmp_path):
    # Временной край покрыт САМОЙ функцией: она не смотрит на
    # существование файла вообще (см. её докстринг) -- вердикт
    # одинаков что до, что после создания файла на диске.
    monkeypatch.setenv("GATEWAY_DB_PATH", str(tmp_path / "fake.db"))
    real_default = tus._default_xlsx_path()
    assert tus._is_safe_to_write(real_default) is False


def test_guard_not_defeated_by_patching_mutable_xlsx_path_global(monkeypatch, tmp_path):
    # ПИН БЛОКЕРА 1 (критик, живая проба 2026-08-05, ДОСЛОВНО): патч
    # МУТИРУЕМОГО tus.XLSX_PATH (ровно тот приём, что раньше несла
    # autouse-фикстура этого файла для СВОЕЙ изоляции) БОЛЬШЕ НЕ снимает
    # защиту с РЕАЛЬНОГО дефолтного пути -- _default_xlsx_path()
    # игнорирует глобал, вычисляет путь заново от __file__ при каждом
    # вызове. До фикса эта проверка была бы True (guard выключен);
    # после -- обязана остаться False.
    monkeypatch.setenv("GATEWAY_DB_PATH", str(tmp_path / "fake.db"))
    monkeypatch.setattr(tus, "XLSX_PATH", tmp_path / "decoy_would_be_default.xlsx")
    real_default = tus._default_xlsx_path()
    assert real_default != tus.XLSX_PATH  # патч действительно разошёлся с эталоном
    assert tus._is_safe_to_write(real_default) is False


def test_guard_allows_explicit_output_elsewhere_positive_control(isolated_db, tmp_path):
    # Позитивный контроль: та же подменённая БД + ЯВНЫЙ --output во
    # временный файл (НЕ дефолтный путь) -- запись состоялась (данные
    # есть -> файл создан). main() вызывается ТОЛЬКО с --output.
    _insert_row(isolated_db, "2020-01-20T12:00:00Z", session_id="s-1")
    out = tmp_path / "alt_output.xlsx"
    code = tus.main(["--output", str(out)])
    assert code == 0
    assert out.exists()


def test_guard_allows_when_gateway_db_path_explicitly_equals_default(monkeypatch):
    # "GATEWAY_DB_PATH выставлен ЯВНО в дефолтный путь репозитория --
    # это НЕ тестовый контекст, запись разрешена" -- сравниваются
    # РАЗРЕШЁННЫЕ АБСОЛЮТНЫЕ пути, не строки. Юнит на _is_safe_to_write
    # напрямую (никакого main()/реального I/O -- не открываем и не
    # трогаем ничего на диске, не создаём и не читаем файл).
    default_db = tus._default_db_path()
    monkeypatch.setenv("GATEWAY_DB_PATH", str(default_db))
    assert tus._is_safe_to_write(tus._default_xlsx_path()) is True


def test_is_safe_to_write_true_when_db_matches_default_and_output_differs(monkeypatch, tmp_path):
    # Граница обратной стороны: БД -- дефолтная (без переопределения),
    # вывод -- НЕ дефолтный путь -- разрешено (тривиально, но
    # закрывает четвёртый квадрант матрицы (db, output)).
    monkeypatch.delenv("GATEWAY_DB_PATH", raising=False)
    assert tus._is_safe_to_write(tmp_path / "anything.xlsx") is True


def test_is_safe_to_write_true_when_db_swapped_and_output_differs(monkeypatch, tmp_path):
    monkeypatch.setenv("GATEWAY_DB_PATH", str(tmp_path / "other.db"))
    assert tus._is_safe_to_write(tmp_path / "anything.xlsx") is True
