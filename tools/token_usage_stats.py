# -*- coding: utf-8 -*-
"""Monthly token-usage statistics -> logs/token_usage.xlsx (append-only).

Monthly window is anchored on the 14th: a window is [14.M, 14.M+1) in
LOCAL time. A window enters the table only once COMPLETED (now >=
window end); re-running is idempotent -- periods already present in a
sheet are never appended twice, and a skipped month is healed by the
next run (every completed missing window is appended, not only the
last one). The earliest data in cc_usage starts 2026-06-13T19:06Z
(21:06 local), so the first window 14.05-14.06.2026 is inherently
partial (one evening); it is included by operator decision 2026-08-04
("если есть более старые данные -- добавь").

Two sheets:
  usage    -- per model family (Haiku/Sonnet/Opus/Fable) per period
  projects -- per project directory (~/.claude/projects/<name>) per period

Accounting is NOT a third ledger: transcript discovery, dedupe
(sessionId:requestId), the cc_usage sqlite import, the price table
(PRICES_PER_TOKEN_USD) and the cost formula are all reused from
tools/usage_report.py (subscription-contour accounting, SIBLING_MAP
axis 2). A model absent from that price table yields cost "н/д" --
never a silent $0 (Rule #1); the fix for such a model is a price row
in usage_report.py, not here.

cc_usage.ts is UTC with 'Z' (axis-2 ts-clock subclass) -- window
bucketing converts to local time before comparing dates.

Scheduled run: Windows Task Scheduler, monthly on the 14th (task
"OS-LLM token usage monthly"), command:
  <python> tools/token_usage_stats.py
The xlsx lands in the repo working tree; it is committed by the next
session's normal batch (axis 9: acceptance vs durability -- the
catch-up logic above makes a lost/uncommitted file recoverable from
transcripts/cc_usage at any time).

OUTPUT PATH OVERRIDE + FAIL-CLOSED GUARD (батч мелочей после
калибровки №6, D-0081, п.1 -- инцидент 2026-08-04 16:52: рабочая
копия logs/token_usage.xlsx была перезаписана неизвестным писателем
с сигнатурой "пересборка со свежей БД", период 14.05 исчез).
Механизм: db_path() (импортирован из usage_report) уважает
GATEWAY_DB_PATH, но XLSX_PATH ниже была НЕИЗМЕНЯЕМОЙ модульной
константой от __file__ -- ВСЕГДА боевой файл репозитория. Тестовый
контекст с подменённой БД, импортировавший модуль и вызвавший
main(), писал в БОЕВОЙ xlsx.

ЕДИНСТВЕННЫЙ механизм переопределения пути вывода -- CLI-флаг
`--output PATH` (не переменная окружения: один механизм, выбор
задокументирован здесь, не угадан молча). Дефолт БЕЗ флага --
прежний XLSX_PATH (logs/token_usage.xlsx) -- штатный запуск
планировщика не меняется.

FAIL-CLOSED GUARD (_is_safe_to_write): если РАЗРЕШЁННЫЙ (resolved)
путь db_path() НЕ равен разрешённому пути дефолтной БД репозитория
(gateway/requests.db от __file__, БЕЗ учёта GATEWAY_DB_PATH -- см.
_default_db_path()), А разрешённый путь вывода равен разрешённому
XLSX_PATH -- main() ОТКАЗЫВАЕТСЯ писать: печатает причину в stderr
и возвращает 1, ДО любого обращения к open_or_create_workbook()
(гвард -- самая первая проверка main(), раньше даже импорта
транскриптов) -- отказ происходит одинаково что при отсутствующем,
что при существующем xlsx (файл никогда не открывается и не
создаётся в обеих ветках). Сравнение -- по РАЗРЕШЁННЫМ АБСОЛЮТНЫМ
ПУТЯМ (Path.resolve()), не по сырым строкам: GATEWAY_DB_PATH,
явно выставленный в тот же дефолтный путь репозитория, -- НЕ
тестовый контекст (пути совпадут после resolve()), запись
разрешена. Обойти guard в законном альтернативном/тестовом прогоне
можно только явным --output, указывающим на ДРУГОЙ (не совпадающий
после resolve() с XLSX_PATH) путь -- сам факт передачи флага
значения не имеет, имеет значение только итоговый разрешённый путь
(--output, буквально равный XLSX_PATH, guard НЕ обходит -- это
единообразнее и безопаснее спец-случая "флаг был передан").
"""
import argparse
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
from usage_report import (  # noqa: E402
    CACHE_WRITE_MULTIPLIER,
    CACHE_READ_MULTIPLIER,
    accounted_cost,
    db_path,
    _connect,
    backfill_costs,
    import_transcripts,
    transcript_glob,
)

XLSX_PATH = Path(__file__).parent.parent / "logs" / "token_usage.xlsx"
USAGE_SHEET = "usage"
PROJECTS_SHEET = "projects"
ANCHOR_DAY = 14

FAMILY_ORDER = ["Haiku", "Sonnet", "Opus", "Fable"]
TOTAL_LABEL = "ИТОГО"

METHODOLOGY = [
    "Статистика использования токенов Claude Code — все проекты машины (~/.claude/projects)",
    "Как считалось:",
    "— Источник: локальные транскрипты сессий; импорт и дедупликация (sessionId:requestId), синтетические записи исключены — через tools/usage_report.py (единый учёт субскрипционного контура).",
    "— Периоды: месячные окна с 14-го по 14-е, локальное время; окно записывается после своего завершения. Пропущенный запуск лечится следующим (дописываются все недостающие завершённые окна).",
    "— Цены — API-тарифы за 1 млн токенов из PRICES_PER_TOKEN_USD (usage_report.py): Haiku 4.5 $1/$5, Sonnet $3/$15, Opus $5/$25, Fable $10/$50 (input/output).",
    f"— Стоимость = input×Pin + output×Pout + cache_write×Pin×{CACHE_WRITE_MULTIPLIER} + cache_read×Pin×{CACHE_READ_MULTIPLIER}. Это оценка «сколько стоил бы тот же объём по API-ценам», а не биллинг подписки.",
    "— Модель без известной цены даёт «н/д» в столбце стоимости (никогда молчаливый $0); фикс — строка цены в usage_report.py.",
    "— Сессий — уникальные сессии периода, коснувшиеся модели; одна сессия может использовать несколько моделей, поэтому строка ИТОГО — уникальные сессии периода, не сумма строк выше.",
    "— Cache read, % — доля кэш-чтений в токенах строки; Output/запрос — средний выход на запрос; $/запрос и $/сессию — стоимость строки на её запросы/сессии; Доля стоимости, % — от итога периода.",
    "— Доля использования, % — доля строки в токенах периода (по столбцу «Токены всего»).",
    "— Лист projects — та же статистика в разрезе проектов (каталогов ~/.claude/projects), отсортировано по стоимости.",
    "— Данные начинаются с 2026-06-13 21:06 (локальное); первое окно 14.05.2026-14.06.2026 из-за этого неполное (только вечер 13.06).",
]

USAGE_HEADER = ["Период", "Модель", "Запросов", "Сессий", "Input", "Output",
                "Cache write", "Cache read", "Токены всего", "Cache read, %",
                "Output/запрос", "Стоимость по API, $", "$/запрос", "$/сессию",
                "Доля стоимости, %", "Доля использования, %"]
USAGE_WIDTHS = [22, 12, 10, 9, 12, 12, 14, 16, 16, 13, 14, 18, 10, 10, 16, 19]
# columns C..I and K hold token/request counts -> #,##0
USAGE_INT_COLS = range(3, 9 + 1)

PROJECTS_HEADER = ["Период", "Проект", "Запросов", "Сессий", "Input", "Output",
                   "Cache write", "Cache read", "Токены всего",
                   "Стоимость по API, $", "Доля стоимости, %",
                   "Доля использования, %"]
PROJECTS_WIDTHS = [22, 48, 10, 9, 12, 12, 14, 16, 16, 18, 16, 19]
PROJECTS_INT_COLS = range(3, 9 + 1)


def family(model: str) -> str:
    m = (model or "").lower()
    if "fable" in m or "mythos" in m:
        return "Fable"
    if "opus" in m:
        return "Opus"
    if "sonnet" in m:
        return "Sonnet"
    if "haiku" in m:
        return "Haiku"
    return model or "unknown"


def parse_ts_local(s: str):
    try:
        return datetime.fromisoformat(s.replace("Z", "+00:00")).astimezone()
    except (ValueError, AttributeError):
        return None


def window_of(dt):
    """(start_y, start_m) of the [14.M, 14.M+1) window containing dt (local)."""
    if dt.day >= ANCHOR_DAY:
        return dt.year, dt.month
    return (dt.year - 1, 12) if dt.month == 1 else (dt.year, dt.month - 1)


def window_label(wy, wm):
    ey, em = (wy + 1, 1) if wm == 12 else (wy, wm + 1)
    return f"14.{wm:02d}.{wy}-14.{em:02d}.{ey}"


def window_end_dt(wy, wm):
    ey, em = (wy + 1, 1) if wm == 12 else (wy, wm + 1)
    return datetime(ey, em, ANCHOR_DAY).astimezone()


def _new_group():
    return {"models": defaultdict(lambda: defaultdict(int)), "sessions": set()}


def aggregate(conn):
    """Returns (by_family, by_project, period_sessions):
    by_family[(wy,wm)][family] = {"models": {model: counters}, "sessions": set}
    by_project[(wy,wm)][project] = same shape
    period_sessions[(wy,wm)] = set of session ids
    """
    by_family = defaultdict(lambda: defaultdict(_new_group))
    by_project = defaultdict(lambda: defaultdict(_new_group))
    period_sessions = defaultdict(set)
    rows = conn.execute(
        "SELECT ts, model, project, session_id, input_tokens, output_tokens,"
        " cache_creation_tokens, cache_read_tokens FROM cc_usage"
    )
    for ts, model, project, session_id, inp, out, cw, cr in rows:
        dt = parse_ts_local(ts)
        if dt is None or not model:
            continue
        wkey = window_of(dt)
        period_sessions[wkey].add(session_id)
        for group in (by_family[wkey][family(model)],
                      by_project[wkey][project or "unknown"]):
            a = group["models"][model]
            a["requests"] += 1
            a["input"] += inp or 0
            a["output"] += out or 0
            a["cache_write"] += cw or 0
            a["cache_read"] += cr or 0
            group["sessions"].add(session_id)
    return by_family, by_project, period_sessions


def _sum_group(group):
    """Collapses one {models, sessions} group -> (counters, cost, unpriced)."""
    s = defaultdict(int)
    cost, unpriced = 0.0, []
    for model, a in group["models"].items():
        for k in ("requests", "input", "output", "cache_write", "cache_read"):
            s[k] += a[k]
        c, _w = accounted_cost(model, a["input"], a["output"],
                               a["cache_write"], a["cache_read"])
        if c is None:
            unpriced.append(model)
        else:
            cost += c
    s["tokens"] = s["input"] + s["output"] + s["cache_write"] + s["cache_read"]
    s["sessions"] = len(group["sessions"])
    return s, cost, unpriced


def _ratio(num, den, digits):
    return round(num / den, digits) if den else 0


def usage_rows(label, groups, period_session_count):
    """Rows for the usage sheet: one per family (known order first) + total."""
    names = [f for f in FAMILY_ORDER if f in groups] + \
            [f for f in sorted(groups) if f not in FAMILY_ORDER]
    summed = {name: _sum_group(groups[name]) for name in names}
    total_cost = sum(cost for _s, cost, unpriced in summed.values() if not unpriced)
    all_priced = all(not unpriced for _s, _c, unpriced in summed.values())
    total_tokens = sum(s["tokens"] for s, _c, _u in summed.values())

    rows = []
    total = defaultdict(int)
    for name in names:
        s, cost, unpriced = summed[name]
        for k in ("requests", "input", "output", "cache_write", "cache_read", "tokens"):
            total[k] += s[k]
        if unpriced:
            cost_c = "н/д (нет цены: %s)" % ", ".join(unpriced)
            per_req = per_sess = share = "н/д"
        else:
            cost_c = round(cost, 2)
            per_req = _ratio(cost, s["requests"], 4)
            per_sess = _ratio(cost, s["sessions"], 2)
            share = _ratio(100 * cost, total_cost, 1) if total_cost else 0
        rows.append([label, name, s["requests"], s["sessions"], s["input"],
                     s["output"], s["cache_write"], s["cache_read"], s["tokens"],
                     _ratio(100 * s["cache_read"], s["tokens"], 1),
                     _ratio(s["output"], s["requests"], 0), cost_c, per_req,
                     per_sess, share,
                     _ratio(100 * s["tokens"], total_tokens, 1)])
    if all_priced:
        cost_c = round(total_cost, 2)
        per_req = _ratio(total_cost, total["requests"], 4)
        per_sess = _ratio(total_cost, period_session_count, 2)
        share = 100.0
    else:
        cost_c = f">= {round(total_cost, 2)} (есть модели без цены)"
        per_req = per_sess = share = "н/д"
    rows.append([label, TOTAL_LABEL, total["requests"], period_session_count,
                 total["input"], total["output"], total["cache_write"],
                 total["cache_read"], total["tokens"],
                 _ratio(100 * total["cache_read"], total["tokens"], 1),
                 _ratio(total["output"], total["requests"], 0), cost_c, per_req,
                 per_sess, share, 100.0])
    return rows


def project_rows(label, groups, period_session_count):
    """Rows for the projects sheet, sorted by cost desc, + total."""
    summed = {name: _sum_group(g) for name, g in groups.items()}
    total_cost = sum(cost for _s, cost, unpriced in summed.values() if not unpriced)
    all_priced = all(not unpriced for _s, _c, unpriced in summed.values())
    total_tokens = sum(s["tokens"] for s, _c, _u in summed.values())

    def sort_key(item):
        _name, (_s, cost, _u) = item
        return -cost
    rows = []
    total = defaultdict(int)
    for name, (s, cost, unpriced) in sorted(summed.items(), key=sort_key):
        for k in ("requests", "input", "output", "cache_write", "cache_read", "tokens"):
            total[k] += s[k]
        if unpriced:
            cost_c = "н/д (нет цены: %s)" % ", ".join(unpriced)
            share = "н/д"
        else:
            cost_c = round(cost, 2)
            share = _ratio(100 * cost, total_cost, 1) if total_cost else 0
        rows.append([label, name, s["requests"], s["sessions"], s["input"],
                     s["output"], s["cache_write"], s["cache_read"], s["tokens"],
                     cost_c, share,
                     _ratio(100 * s["tokens"], total_tokens, 1)])
    rows.append([label, TOTAL_LABEL, total["requests"], period_session_count,
                 total["input"], total["output"], total["cache_write"],
                 total["cache_read"], total["tokens"],
                 round(total_cost, 2) if all_priced else f">= {round(total_cost, 2)} (есть модели без цены)",
                 100.0 if all_priced else "н/д", 100.0])
    return rows


def _init_sheet(ws, header, widths, methodology=None):
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    if methodology:
        for line in methodology:
            ws.append([line])
        ws[1][0].font = Font(bold=True, size=12)
        ws.append([])
    ws.append(header)
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def open_or_create_workbook(output_path: Path = None):
    from openpyxl import Workbook, load_workbook
    output_path = output_path or _default_xlsx_path()
    if output_path.exists():
        wb = load_workbook(output_path)
        return wb, wb[USAGE_SHEET], wb[PROJECTS_SHEET]
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws_u = wb.active
    ws_u.title = USAGE_SHEET
    _init_sheet(ws_u, USAGE_HEADER, USAGE_WIDTHS, METHODOLOGY)
    ws_p = wb.create_sheet(PROJECTS_SHEET)
    _init_sheet(ws_p, PROJECTS_HEADER, PROJECTS_WIDTHS,
                ["Разрез по проектам (каталоги ~/.claude/projects); методика — на листе usage."])
    return wb, ws_u, ws_p


def existing_periods(ws):
    return {row[0].value for row in ws.iter_rows(min_row=1) if row[0].value}


def _append_rows(ws, rows, int_cols):
    from openpyxl.styles import Font
    for row in rows:
        ws.append(row)
        if row[1] == TOTAL_LABEL:
            for cell in ws[ws.max_row]:
                cell.font = Font(bold=True)
        for idx in int_cols:
            ws[ws.max_row][idx - 1].number_format = "#,##0"


def _default_db_path() -> Path:
    """Дефолтная БД репозитория, БЕЗ учёта GATEWAY_DB_PATH -- эталон, с
    которым сравнивается db_path() (которая env-переменную учитывает)
    в guard'е ниже. См. докстринг модуля "FAIL-CLOSED GUARD"."""
    return Path(__file__).parent.parent / "gateway" / "requests.db"


def _default_xlsx_path() -> Path:
    """Дефолтный боевой xlsx, ВСЕГДА свежевычисленный от __file__ --
    симметрично _default_db_path(), НЕ читает модульную переменную
    XLSX_PATH. ПОЧЕМУ (пересдача, блокер 1 критика, живая проба
    2026-08-05): guard и выбор дефолтного пути вывода раньше читали
    XLSX_PATH -- ИЗМЕНЯЕМЫЙ модульный глобал. Контекст, патчащий
    XLSX_PATH ради ИЗОЛЯЦИИ СВОИХ путей вывода (ровно то, что делала
    autouse-фикстура старой версии test_token_usage_stats.py), тем
    самым переопределял и то, ЧТО ГВАРД СЧИТАЕТ "боевым путём" --
    патч глобала в одном контексте снимал защиту с РЕАЛЬНОГО файла
    для explicit --output, указывающего буквально на него (живая
    проба критика: guard() -> True после патча XLSX_PATH, main(
    --output <живой logs/token_usage.xlsx>) реально дописал
    выдуманный период, 9172->9456 байт). Guard/дефолт вывода ниже
    ТЕПЕРЬ читают ЭТУ функцию (не глобал) -- патч XLSX_PATH (или
    любого другого атрибута модуля) больше НЕ влияет на то, что
    считается "боевым" файлом."""
    return Path(__file__).parent.parent / "logs" / "token_usage.xlsx"


def _is_safe_to_write(output_path: Path) -> bool:
    """False -- ТОЛЬКО когда db_path() (уважает GATEWAY_DB_PATH)
    указывает НЕ на дефолтную БД репозитория, А output_path
    разрешается в тот же путь, что _default_xlsx_path() (боевой файл,
    ИММУТАБЕЛЬНЫЙ эталон от __file__ -- см. её докстринг за полное
    обоснование симметрии с _default_db_path(), пересдача блокера 1).
    Сравнение -- по Path.resolve(), не по сырым строкам (см. докстринг
    модуля). resolve() на несуществующем пути безопасен на
    Windows/POSIX (не требует существования файла)."""
    current_db = db_path().resolve()
    default_db = _default_db_path().resolve()
    out = output_path.resolve()
    default_out = _default_xlsx_path().resolve()
    if current_db != default_db and out == default_out:
        return False
    return True


def main(argv=None) -> int:
    parser = argparse.ArgumentParser(
        description="Monthly token-usage statistics -> xlsx (append-only)."
    )
    parser.add_argument(
        "--output",
        default=None,
        help=(
            "Путь вывода xlsx (по умолчанию logs/token_usage.xlsx). "
            "ЕДИНСТВЕННЫЙ механизм переопределения пути -- см. докстринг "
            "модуля 'OUTPUT PATH OVERRIDE + FAIL-CLOSED GUARD'."
        ),
    )
    args = parser.parse_args(argv)
    output_path = Path(args.output) if args.output else _default_xlsx_path()

    if not _is_safe_to_write(output_path):
        print(
            "REFUSED: GATEWAY_DB_PATH указывает НЕ на дефолтную БД "
            f"репозитория ({_default_db_path()}), а путь вывода -- "
            f"дефолтный боевой файл ({_default_xlsx_path()}) -- guard против "
            "записи тестовым/подменённым контекстом в боевой xlsx (батч "
            "мелочей после калибровки №6, D-0081, п.1). Укажи --output "
            "<path> явно, если это осознанный альтернативный прогон.",
            file=sys.stderr,
        )
        return 1

    imported, sessions, warnings = import_transcripts(transcript_glob(), db_path())
    print(f"imported {imported} new rows from {len(sessions)} sessions")
    conn = _connect(db_path())
    try:
        backfilled = backfill_costs(conn)
        conn.commit()
        if backfilled:
            print(f"backfilled costs for {backfilled} rows")
        by_family, by_project, period_sessions = aggregate(conn)
    finally:
        conn.close()
    for w in warnings:
        print(w)

    wb, ws_u, ws_p = open_or_create_workbook(output_path)
    now = datetime.now().astimezone()
    appended = []
    for wkey in sorted(by_family):
        if window_end_dt(*wkey) > now:
            continue  # window not finished yet
        label = window_label(*wkey)
        n_sessions = len(period_sessions[wkey])
        if label not in existing_periods(ws_u):
            _append_rows(ws_u, usage_rows(label, by_family[wkey], n_sessions),
                         USAGE_INT_COLS)
            appended.append(label)
        if label not in existing_periods(ws_p):
            _append_rows(ws_p, project_rows(label, by_project[wkey], n_sessions),
                         PROJECTS_INT_COLS)
    if appended:
        wb.save(output_path)
        print(f"appended periods: {', '.join(appended)} -> {output_path}")
    else:
        print("no new completed periods to append")
    return 0


if __name__ == "__main__":
    sys.exit(main())
