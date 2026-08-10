"""Tests for tools/token_usage_stats.py. Isolation pattern -- same as
tools/test_usage_report.py: a temporary sqlite DB (created via a direct
INSERT into cc_usage through usage_report.SCHEMA, with NO transcripts),
GATEWAY_DB_PATH points at it (motive: see the "OUTPUT PATH OVERRIDE +
FAIL-CLOSED GUARD" section of token_usage_stats.py's own docstring).

transcript_glob is monkeypatched to an empty temp directory --
import_transcripts() never scans the real ~/.claude/projects tree.

GUARD ISOLATION RULE: this file does NOT patch
tools/token_usage_stats.XLSX_PATH (no autouse fixture on the module
global) -- that exact trick (patching the mutable XLSX_PATH for one's
own isolation) is what used to disarm the guard on the REAL live file
for any call with an explicit --output pointing literally at it (see
_default_xlsx_path()'s docstring in token_usage_stats.py for the full
account). The rule is simple: EVERY call to tus.main() in this file
carries an EXPLICIT `--output <tmp path>`; the guard's behavior on the
REAL default path is checked ONLY via a direct call to the pure
function `_is_safe_to_write()` (a verdict, not a write -- see the
FAIL-CLOSED GUARD section below) -- no test in this file writes or
reads the repo's real logs/token_usage.xlsx in any form.

Run from tools/: python -m pytest test_token_usage_stats.py -q
"""

import sqlite3
from datetime import datetime, timezone
from pathlib import Path

import pytest

import token_usage_stats as tus
import usage_report

FAMILY_SONNET = "claude-sonnet-5"


def _make_db(tmp_path) -> Path:
    db_file = tmp_path / "requests.db"
    conn = sqlite3.connect(db_file)
    conn.execute(usage_report.SCHEMA)
    conn.commit()
    conn.close()
    return db_file


def _insert_row(db_file, ts, model=FAMILY_SONNET, project="projA",
                 session_id="s-1", dedupe_suffix=None,
                 input_tokens=100, output_tokens=50):
    cost, _ = usage_report.accounted_cost(model, input_tokens, output_tokens, 0, 0)
    dedupe = f"{session_id}:{dedupe_suffix or ts}"
    conn = sqlite3.connect(db_file)
    conn.execute(
        """
        INSERT INTO cc_usage
            (ts, project, session_id, turn_index, model, input_tokens,
             output_tokens, cache_creation_tokens, cache_read_tokens,
             accounted_cost_usd, traffic_kind, is_sidechain, dedupe_key)
        VALUES (?, ?, ?, 0, ?, ?, ?, 0, 0, ?, 'real', 0, ?)
        """,
        (ts, project, session_id, model, input_tokens, output_tokens, cost, dedupe),
    )
    conn.commit()
    conn.close()


@pytest.fixture()
def isolated_db(tmp_path, monkeypatch):
    """An isolated DB (see the module docstring) + an empty
    transcript_glob -- zero transcript scanning of the real
    ~/.claude/projects tree."""
    db_file = _make_db(tmp_path)
    monkeypatch.setenv("GATEWAY_DB_PATH", str(db_file))
    empty_dir = tmp_path / "no_transcripts"
    empty_dir.mkdir()
    monkeypatch.setattr(
        tus, "transcript_glob",
        lambda base_dir=None: [str(empty_dir / "*" / "*.jsonl")],
    )
    return db_file


# ---------------------------------------------------------------------
# Window bucketing at the 14th boundary (rule 6a -- both sides).
# ---------------------------------------------------------------------


def test_window_of_exact_local_midnight_14th_is_new_window():
    # "a row exactly on the boundary (the 14th, 00:00:00 local time)
    # belongs to the NEW window" -- a literal boundary test.
    assert tus.window_of(datetime(2026, 6, 14, 0, 0, 0)) == (2026, 6)


def test_window_of_one_second_before_midnight_14th_is_old_window():
    assert tus.window_of(datetime(2026, 6, 13, 23, 59, 59)) == (2026, 5)


def test_utc_stored_ts_buckets_by_local_time_not_utc_date(isolated_db):
    # cc_usage.ts is stored in UTC ('Z'); the window is bucketed by
    # LOCAL time. Noon UTC ("T12:00:00Z") converts to local time in
    # ANY realistic timezone (-12..+14h) without crossing the window
    # boundary in the OPPOSITE direction -- the test is robust to the
    # run machine's TZ (it does not hardcode a specific offset).
    _insert_row(isolated_db, "2026-06-12T12:00:00Z", session_id="s-old")  # < 14 everywhere
    _insert_row(isolated_db, "2026-06-14T12:00:00Z", session_id="s-new")  # >= 14 everywhere

    conn = sqlite3.connect(isolated_db)
    try:
        by_family, _by_project, _sessions = tus.aggregate(conn)
    finally:
        conn.close()

    windows = set(by_family.keys())
    assert (2026, 5) in windows  # the old window caught the "12th" row
    assert (2026, 6) in windows  # the new window caught the "14th" row


# ---------------------------------------------------------------------
# Append idempotency -- a second run does not duplicate the period on
# EITHER sheet, usage or projects.
# ---------------------------------------------------------------------


def test_append_is_idempotent_on_both_sheets(isolated_db, tmp_path):
    # A completed period -- ts is deliberately in the past (a fully
    # closed 14.01-14.02.2020 window).
    _insert_row(isolated_db, "2020-01-20T12:00:00Z", session_id="s-1")
    out = tmp_path / "out.xlsx"

    code1 = tus.main(["--output", str(out)])
    assert code1 == 0
    assert out.exists()

    from openpyxl import load_workbook
    wb1 = load_workbook(out)
    usage_rows_1 = wb1[tus.USAGE_SHEET].max_row
    projects_rows_1 = wb1[tus.PROJECTS_SHEET].max_row

    code2 = tus.main(["--output", str(out)])
    assert code2 == 0

    wb2 = load_workbook(out)
    usage_rows_2 = wb2[tus.USAGE_SHEET].max_row
    projects_rows_2 = wb2[tus.PROJECTS_SHEET].max_row

    assert usage_rows_2 == usage_rows_1
    assert projects_rows_2 == projects_rows_1
    # the period is not duplicated -- the number of ROWS carrying this
    # label (family rows + TOTAL, wb1 already gives the reference N for
    # one occurrence of the period) does not change between runs.
    label = tus.window_label(2020, 1)
    label_rows_1 = [
        row[0].value for row in wb1[tus.USAGE_SHEET].iter_rows(min_row=1)
        if row[0].value == label
    ]
    label_rows_2 = [
        row[0].value for row in wb2[tus.USAGE_SHEET].iter_rows(min_row=1)
        if row[0].value == label
    ]
    assert len(label_rows_1) >= 1
    assert len(label_rows_2) == len(label_rows_1)
    proj_label_rows_1 = [
        row[0].value for row in wb1[tus.PROJECTS_SHEET].iter_rows(min_row=1)
        if row[0].value == label
    ]
    proj_label_rows_2 = [
        row[0].value for row in wb2[tus.PROJECTS_SHEET].iter_rows(min_row=1)
        if row[0].value == label
    ]
    assert len(proj_label_rows_1) >= 1
    assert len(proj_label_rows_2) == len(proj_label_rows_1)


# ---------------------------------------------------------------------
# An unfinished (current) window is NOT appended.
# ---------------------------------------------------------------------


def test_current_incomplete_window_not_appended(isolated_db, tmp_path):
    now_utc_ts = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    _insert_row(isolated_db, now_utc_ts, session_id="s-now")
    out = tmp_path / "out.xlsx"
    code = tus.main(["--output", str(out)])
    assert code == 0
    # zero completed periods -- the file is NOT created.
    assert not out.exists()


def test_completed_and_incomplete_windows_together_only_completed_appended(isolated_db, tmp_path):
    _insert_row(isolated_db, "2020-01-20T12:00:00Z", session_id="s-old", dedupe_suffix="a")
    now_utc_ts = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    _insert_row(isolated_db, now_utc_ts, session_id="s-now", dedupe_suffix="b")
    out = tmp_path / "out.xlsx"
    code = tus.main(["--output", str(out)])
    assert code == 0
    assert out.exists()

    from openpyxl import load_workbook
    wb = load_workbook(out)
    periods = {row[0].value for row in wb[tus.USAGE_SHEET].iter_rows(min_row=1) if row[0].value}
    assert tus.window_label(2020, 1) in periods
    # the current (incomplete) window did not appear on the sheet
    wkey_now = tus.window_of(datetime.now())
    assert tus.window_label(*wkey_now) not in periods


# ---------------------------------------------------------------------
# An empty DB / zero completed periods.
# ---------------------------------------------------------------------


def test_empty_db_no_file_created_and_message_printed(isolated_db, tmp_path, capsys):
    out = tmp_path / "out.xlsx"
    code = tus.main(["--output", str(out)])
    assert code == 0
    assert not out.exists()
    captured = capsys.readouterr()
    assert "no new completed periods to append" in captured.out


# ---------------------------------------------------------------------
# FAIL-CLOSED GUARD. The guard's behavior on the REAL default path is
# proved ONLY by the verdict of the pure function _is_safe_to_write()
# -- no test calls main() with no --output, and none creates/opens a
# file at the real default path (see the module docstring above,
# "GUARD ISOLATION RULE").
# ---------------------------------------------------------------------


def test_is_safe_to_write_blocks_real_default_path_when_db_swapped(monkeypatch, tmp_path):
    # The temporal edge is covered by the function ITSELF: it never
    # looks at whether the file exists at all (see its docstring) --
    # the verdict is the same before and after the file is created on
    # disk.
    monkeypatch.setenv("GATEWAY_DB_PATH", str(tmp_path / "fake.db"))
    real_default = tus._default_xlsx_path()
    assert tus._is_safe_to_write(real_default) is False


def test_guard_not_defeated_by_patching_mutable_xlsx_path_global(monkeypatch, tmp_path):
    # PIN of the finding, VERBATIM: patching the MUTABLE tus.XLSX_PATH
    # (exactly the trick this file's own autouse fixture used to carry
    # for ITS OWN isolation) no longer disarms protection of the REAL
    # default path -- _default_xlsx_path() ignores the global,
    # recomputing the path fresh from __file__ on every call. Before
    # the fix this check would have been True (guard disabled); after
    # the fix it must stay False.
    monkeypatch.setenv("GATEWAY_DB_PATH", str(tmp_path / "fake.db"))
    monkeypatch.setattr(tus, "XLSX_PATH", tmp_path / "decoy_would_be_default.xlsx")
    real_default = tus._default_xlsx_path()
    assert real_default != tus.XLSX_PATH  # the patch really did diverge from the reference
    assert tus._is_safe_to_write(real_default) is False


def test_guard_allows_explicit_output_elsewhere_positive_control(isolated_db, tmp_path):
    # Positive control: the same swapped DB + an EXPLICIT --output to a
    # temp file (NOT the default path) -- the write goes through (data
    # exists -> the file is created). main() is called ONLY with
    # --output.
    _insert_row(isolated_db, "2020-01-20T12:00:00Z", session_id="s-1")
    out = tmp_path / "alt_output.xlsx"
    code = tus.main(["--output", str(out)])
    assert code == 0
    assert out.exists()


def test_guard_allows_when_gateway_db_path_explicitly_equals_default(monkeypatch):
    # "GATEWAY_DB_PATH explicitly set to the repo's default path -- this
    # is NOT a test context, the write is allowed" -- the RESOLVED
    # ABSOLUTE paths are compared, not strings. A unit test directly on
    # _is_safe_to_write (no main()/real I/O -- nothing on disk is
    # opened or touched, no file is created or read).
    default_db = tus._default_db_path()
    monkeypatch.setenv("GATEWAY_DB_PATH", str(default_db))
    assert tus._is_safe_to_write(tus._default_xlsx_path()) is True


def test_is_safe_to_write_true_when_db_matches_default_and_output_differs(monkeypatch, tmp_path):
    # The opposite-side boundary: the DB is default (no override), the
    # output is NOT the default path -- allowed (trivial, but closes
    # the fourth quadrant of the (db, output) matrix).
    monkeypatch.delenv("GATEWAY_DB_PATH", raising=False)
    assert tus._is_safe_to_write(tmp_path / "anything.xlsx") is True


def test_is_safe_to_write_true_when_db_swapped_and_output_differs(monkeypatch, tmp_path):
    monkeypatch.setenv("GATEWAY_DB_PATH", str(tmp_path / "other.db"))
    assert tus._is_safe_to_write(tmp_path / "anything.xlsx") is True
