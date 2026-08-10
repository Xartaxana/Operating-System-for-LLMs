# -*- coding: utf-8 -*-
"""Monthly token-usage statistics -> logs/token_usage.xlsx (append-only).

Monthly window is anchored on the 14th: a window is [14.M, 14.M+1) in
LOCAL time. A window enters the table only once COMPLETED (now >=
window end); re-running is idempotent -- periods already present in a
sheet are never appended twice, and a skipped month is healed by the
next run (every completed missing window is appended, not only the
last one). A window whose start predates the earliest data actually
present in cc_usage is inherently partial (covers however much of the
window has data); that is expected, not an error.

Two sheets:
  usage    -- per model family (Haiku/Sonnet/Opus/Fable) per period
  projects -- per project directory (~/.claude/projects/<name>) per period

Accounting is NOT a third ledger: transcript discovery, dedupe
(sessionId:requestId), the cc_usage sqlite import, the price table
(PRICES_PER_TOKEN_USD) and the cost formula are all reused from
tools/usage_report.py (subscription-contour accounting, one source of
truth). A model absent from that price table yields cost "n/a" --
never a silent $0 (fail loud, not silent); the fix for such a model is
a price row in usage_report.py, not here.

cc_usage.ts is UTC with 'Z' -- window bucketing converts to local time
before comparing dates.

Scheduled run: an OS-level scheduled task, monthly on the 14th,
command:
  <python> tools/token_usage_stats.py
The xlsx lands in the repo working tree; it is committed by the next
session's normal batch (the catch-up logic above makes a lost/
uncommitted file recoverable from transcripts/cc_usage at any time).

OUTPUT PATH OVERRIDE + FAIL-CLOSED GUARD.
Motivating incident (own inline account): a test run swapped the
accounting database via an environment variable (GATEWAY_DB_PATH) so
it would read synthetic data, then ran this module's main() with no
explicit output override -- the module wrote its result into the
LIVE repo xlsx anyway, because the output-path constant it consulted
was resolved from the database override's PRESENCE, not from where
the caller told it to write. The mechanism: db_path() (imported from
usage_report) respects GATEWAY_DB_PATH, but the module's own XLSX_PATH
constant used to be computed once, unconditionally, from __file__ --
always the live repo file, regardless of which database was actually
being read. A test context with a swapped DB, importing the module
and calling main(), silently wrote into the LIVE xlsx.

The ONLY mechanism for overriding the output path is the CLI flag
`--output PATH` (not an environment variable: a single mechanism, the
choice is documented here, not guessed silently). The default with no
flag is the same XLSX_PATH as before (logs/token_usage.xlsx) -- the
normal scheduled run is unchanged.

FAIL-CLOSED GUARD (_is_safe_to_write): if the RESOLVED path of
db_path() is NOT equal to the resolved path of the repo's default
database (gateway/requests.db from __file__, ignoring GATEWAY_DB_PATH
-- see _default_db_path()), AND the resolved output path equals the
resolved default XLSX_PATH -- main() REFUSES to write: it prints the
reason to stderr and returns 1, BEFORE touching
open_or_create_workbook() (the guard is the very first check in
main(), before even importing transcripts) -- the refusal is the same
whether the xlsx already exists or not (the file is never opened or
created on either branch). The comparison is on RESOLVED ABSOLUTE
PATHS (Path.resolve()), not raw strings: GATEWAY_DB_PATH explicitly
set to the same default repo path is NOT a test context (the paths
match after resolve()), and the write is allowed. The only legitimate
way to bypass the guard in an alternative/test run is an explicit
--output pointing at a path that DIFFERS (after resolve()) from
XLSX_PATH -- passing the flag at all has no special effect by itself,
only the final resolved path matters (an --output literally equal to
XLSX_PATH does NOT bypass the guard -- this is simpler and safer than
special-casing "the flag was passed").

A second finding from a live probe of this same guard: an EARLIER
version of this guard, and of the default-output-path helper, both
read the module-level XLSX_PATH global directly. A caller that patches
that mutable global for its OWN test isolation (redirecting where ITS
OWN writes land) thereby also redefines what the GUARD itself
considers "the live path" -- patching the global in one test context
silently disarmed protection of the REAL file for any other call in
that same process passing an explicit --output that points literally
at it. The fix: the guard and the default-output helper below now
read a dedicated function (_default_xlsx_path(), always recomputed
from __file__), never the mutable module global -- patching XLSX_PATH
(or any other module attribute) no longer changes what counts as "the
live" file.
"""
import argparse
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

_TOOLS_DIR = str(Path(__file__).resolve().parent)
if _TOOLS_DIR not in sys.path:
    sys.path.insert(0, _TOOLS_DIR)
from usage_report import (  # noqa: E402
    CACHE_WRITE_MULTIPLIER,
    CACHE_READ_MULTIPLIER,
    accounted_cost,
    db_path,
    _connect,
    backfill_costs,
    import_transcripts,
    transcript_glob,
)

XLSX_PATH = Path(__file__).parent.parent / "logs" / "token_usage.xlsx"
USAGE_SHEET = "usage"
PROJECTS_SHEET = "projects"
ANCHOR_DAY = 14

FAMILY_ORDER = ["Haiku", "Sonnet", "Opus", "Fable"]
TOTAL_LABEL = "TOTAL"

METHODOLOGY = [
    "Claude Code token usage statistics -- all projects on this machine (~/.claude/projects)",
    "How it's computed:",
    "-- Source: local session transcripts; import and dedupe (sessionId:requestId), synthetic entries excluded -- via tools/usage_report.py (single accounting path for the subscription contour).",
    "-- Periods: monthly windows from the 14th to the 14th, local time; a window is recorded once it has completed. A missed run is healed by the next one (every missing completed window is appended).",
    "-- Prices -- API rates per 1M tokens from PRICES_PER_TOKEN_USD (usage_report.py); see that table for the current per-model rates (input/output).",
    f"-- Cost = input*Pin + output*Pout + cache_write*Pin*{CACHE_WRITE_MULTIPLIER} + cache_read*Pin*{CACHE_READ_MULTIPLIER}. This is an estimate of \"what the same volume would cost at API prices\", not a subscription bill.",
    "-- A model with no known price gets \"n/a\" in the cost column (never a silent $0); the fix is a price row in usage_report.py.",
    "-- Sessions -- unique sessions in the period that touched the model; one session can use several models, so the TOTAL row is unique sessions for the period, not the sum of the rows above.",
    "-- Cache read, % -- share of cache reads in the row's tokens; Output/request -- average output per request; $/request and $/session -- the row's cost per its requests/sessions; Cost share, % -- of the period's total.",
    "-- Usage share, % -- the row's share of the period's tokens (by the \"Tokens total\" column).",
    "-- The projects sheet -- the same statistics broken down by project (~/.claude/projects directories), sorted by cost.",
]

USAGE_HEADER = ["Period", "Model", "Requests", "Sessions", "Input", "Output",
                "Cache write", "Cache read", "Tokens total", "Cache read, %",
                "Output/request", "Cost by API, $", "$/request", "$/session",
                "Cost share, %", "Usage share, %"]
USAGE_WIDTHS = [22, 12, 10, 9, 12, 12, 14, 16, 16, 13, 14, 18, 10, 10, 16, 19]
# columns C..I and K hold token/request counts -> #,##0
USAGE_INT_COLS = range(3, 9 + 1)

PROJECTS_HEADER = ["Period", "Project", "Requests", "Sessions", "Input", "Output",
                   "Cache write", "Cache read", "Tokens total",
                   "Cost by API, $", "Cost share, %",
                   "Usage share, %"]
PROJECTS_WIDTHS = [22, 48, 10, 9, 12, 12, 14, 16, 16, 18, 16, 19]
PROJECTS_INT_COLS = range(3, 9 + 1)


def family(model: str) -> str:
    m = (model or "").lower()
    if "fable" in m or "mythos" in m:
        return "Fable"
    if "opus" in m:
        return "Opus"
    if "sonnet" in m:
        return "Sonnet"
    if "haiku" in m:
        return "Haiku"
    return model or "unknown"


def parse_ts_local(s: str):
    try:
        return datetime.fromisoformat(s.replace("Z", "+00:00")).astimezone()
    except (ValueError, AttributeError):
        return None


def window_of(dt):
    """(start_y, start_m) of the [14.M, 14.M+1) window containing dt (local)."""
    if dt.day >= ANCHOR_DAY:
        return dt.year, dt.month
    return (dt.year - 1, 12) if dt.month == 1 else (dt.year, dt.month - 1)


def window_label(wy, wm):
    ey, em = (wy + 1, 1) if wm == 12 else (wy, wm + 1)
    return f"14.{wm:02d}.{wy}-14.{em:02d}.{ey}"


def window_end_dt(wy, wm):
    ey, em = (wy + 1, 1) if wm == 12 else (wy, wm + 1)
    return datetime(ey, em, ANCHOR_DAY).astimezone()


def _new_group():
    return {"models": defaultdict(lambda: defaultdict(int)), "sessions": set()}


def aggregate(conn):
    """Returns (by_family, by_project, period_sessions):
    by_family[(wy,wm)][family] = {"models": {model: counters}, "sessions": set}
    by_project[(wy,wm)][project] = same shape
    period_sessions[(wy,wm)] = set of session ids
    """
    by_family = defaultdict(lambda: defaultdict(_new_group))
    by_project = defaultdict(lambda: defaultdict(_new_group))
    period_sessions = defaultdict(set)
    rows = conn.execute(
        "SELECT ts, model, project, session_id, input_tokens, output_tokens,"
        " cache_creation_tokens, cache_read_tokens FROM cc_usage"
    )
    for ts, model, project, session_id, inp, out, cw, cr in rows:
        dt = parse_ts_local(ts)
        if dt is None or not model:
            continue
        wkey = window_of(dt)
        period_sessions[wkey].add(session_id)
        for group in (by_family[wkey][family(model)],
                      by_project[wkey][project or "unknown"]):
            a = group["models"][model]
            a["requests"] += 1
            a["input"] += inp or 0
            a["output"] += out or 0
            a["cache_write"] += cw or 0
            a["cache_read"] += cr or 0
            group["sessions"].add(session_id)
    return by_family, by_project, period_sessions


def _sum_group(group):
    """Collapses one {models, sessions} group -> (counters, cost, unpriced)."""
    s = defaultdict(int)
    cost, unpriced = 0.0, []
    for model, a in group["models"].items():
        for k in ("requests", "input", "output", "cache_write", "cache_read"):
            s[k] += a[k]
        c, _w = accounted_cost(model, a["input"], a["output"],
                               a["cache_write"], a["cache_read"])
        if c is None:
            unpriced.append(model)
        else:
            cost += c
    s["tokens"] = s["input"] + s["output"] + s["cache_write"] + s["cache_read"]
    s["sessions"] = len(group["sessions"])
    return s, cost, unpriced


def _ratio(num, den, digits):
    return round(num / den, digits) if den else 0


def usage_rows(label, groups, period_session_count):
    """Rows for the usage sheet: one per family (known order first) + total."""
    names = [f for f in FAMILY_ORDER if f in groups] + \
            [f for f in sorted(groups) if f not in FAMILY_ORDER]
    summed = {name: _sum_group(groups[name]) for name in names}
    total_cost = sum(cost for _s, cost, unpriced in summed.values() if not unpriced)
    all_priced = all(not unpriced for _s, _c, unpriced in summed.values())
    total_tokens = sum(s["tokens"] for s, _c, _u in summed.values())

    rows = []
    total = defaultdict(int)
    for name in names:
        s, cost, unpriced = summed[name]
        for k in ("requests", "input", "output", "cache_write", "cache_read", "tokens"):
            total[k] += s[k]
        if unpriced:
            cost_c = "n/a (no price for: %s)" % ", ".join(unpriced)
            per_req = per_sess = share = "n/a"
        else:
            cost_c = round(cost, 2)
            per_req = _ratio(cost, s["requests"], 4)
            per_sess = _ratio(cost, s["sessions"], 2)
            share = _ratio(100 * cost, total_cost, 1) if total_cost else 0
        rows.append([label, name, s["requests"], s["sessions"], s["input"],
                     s["output"], s["cache_write"], s["cache_read"], s["tokens"],
                     _ratio(100 * s["cache_read"], s["tokens"], 1),
                     _ratio(s["output"], s["requests"], 0), cost_c, per_req,
                     per_sess, share,
                     _ratio(100 * s["tokens"], total_tokens, 1)])
    if all_priced:
        cost_c = round(total_cost, 2)
        per_req = _ratio(total_cost, total["requests"], 4)
        per_sess = _ratio(total_cost, period_session_count, 2)
        share = 100.0
    else:
        cost_c = f">= {round(total_cost, 2)} (some models have no price)"
        per_req = per_sess = share = "n/a"
    rows.append([label, TOTAL_LABEL, total["requests"], period_session_count,
                 total["input"], total["output"], total["cache_write"],
                 total["cache_read"], total["tokens"],
                 _ratio(100 * total["cache_read"], total["tokens"], 1),
                 _ratio(total["output"], total["requests"], 0), cost_c, per_req,
                 per_sess, share, 100.0])
    return rows


def project_rows(label, groups, period_session_count):
    """Rows for the projects sheet, sorted by cost desc, + total."""
    summed = {name: _sum_group(g) for name, g in groups.items()}
    total_cost = sum(cost for _s, cost, unpriced in summed.values() if not unpriced)
    all_priced = all(not unpriced for _s, _c, unpriced in summed.values())
    total_tokens = sum(s["tokens"] for s, _c, _u in summed.values())

    def sort_key(item):
        _name, (_s, cost, _u) = item
        return -cost
    rows = []
    total = defaultdict(int)
    for name, (s, cost, unpriced) in sorted(summed.items(), key=sort_key):
        for k in ("requests", "input", "output", "cache_write", "cache_read", "tokens"):
            total[k] += s[k]
        if unpriced:
            cost_c = "n/a (no price for: %s)" % ", ".join(unpriced)
            share = "n/a"
        else:
            cost_c = round(cost, 2)
            share = _ratio(100 * cost, total_cost, 1) if total_cost else 0
        rows.append([label, name, s["requests"], s["sessions"], s["input"],
                     s["output"], s["cache_write"], s["cache_read"], s["tokens"],
                     cost_c, share,
                     _ratio(100 * s["tokens"], total_tokens, 1)])
    rows.append([label, TOTAL_LABEL, total["requests"], period_session_count,
                 total["input"], total["output"], total["cache_write"],
                 total["cache_read"], total["tokens"],
                 round(total_cost, 2) if all_priced else f">= {round(total_cost, 2)} (some models have no price)",
                 100.0 if all_priced else "n/a", 100.0])
    return rows


def _init_sheet(ws, header, widths, methodology=None):
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    if methodology:
        for line in methodology:
            ws.append([line])
        ws[1][0].font = Font(bold=True, size=12)
        ws.append([])
    ws.append(header)
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def open_or_create_workbook(output_path: Path = None):
    from openpyxl import Workbook, load_workbook
    output_path = output_path or _default_xlsx_path()
    if output_path.exists():
        wb = load_workbook(output_path)
        return wb, wb[USAGE_SHEET], wb[PROJECTS_SHEET]
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws_u = wb.active
    ws_u.title = USAGE_SHEET
    _init_sheet(ws_u, USAGE_HEADER, USAGE_WIDTHS, METHODOLOGY)
    ws_p = wb.create_sheet(PROJECTS_SHEET)
    _init_sheet(ws_p, PROJECTS_HEADER, PROJECTS_WIDTHS,
                ["Breakdown by project (~/.claude/projects directories); methodology -- see the usage sheet."])
    return wb, ws_u, ws_p


def existing_periods(ws):
    return {row[0].value for row in ws.iter_rows(min_row=1) if row[0].value}


def _append_rows(ws, rows, int_cols):
    from openpyxl.styles import Font
    for row in rows:
        ws.append(row)
        if row[1] == TOTAL_LABEL:
            for cell in ws[ws.max_row]:
                cell.font = Font(bold=True)
        for idx in int_cols:
            ws[ws.max_row][idx - 1].number_format = "#,##0"


def _default_db_path() -> Path:
    """The repo's default database, IGNORING GATEWAY_DB_PATH -- the
    reference _is_safe_to_write() compares db_path() (which DOES
    respect the env override) against, below. See the module docstring
    "FAIL-CLOSED GUARD"."""
    return Path(__file__).parent.parent / "gateway" / "requests.db"


def _default_xlsx_path() -> Path:
    """The default live xlsx, ALWAYS freshly computed from __file__ --
    symmetric with _default_db_path(), never reads the module-level
    XLSX_PATH variable. WHY (see the module docstring's second
    finding): the guard and the default-output-path choice used to
    read XLSX_PATH -- a MUTABLE module global. A context patching
    XLSX_PATH for its OWN output isolation thereby also redefined what
    the GUARD considers "the live path" -- patching the global in one
    context silently disarmed protection of the REAL file for an
    explicit --output that pointed literally at it. The guard/default
    below now read THIS FUNCTION (not the global) -- patching
    XLSX_PATH (or any other module attribute) no longer affects what
    counts as "the live" file."""
    return Path(__file__).parent.parent / "logs" / "token_usage.xlsx"


def _is_safe_to_write(output_path: Path) -> bool:
    """False -- ONLY when db_path() (which respects GATEWAY_DB_PATH)
    points somewhere OTHER than the repo's default database, AND
    output_path resolves to the same path as _default_xlsx_path() (the
    live file, an IMMUTABLE reference computed from __file__ -- see
    its docstring for the full rationale, symmetric with
    _default_db_path()). The comparison is on Path.resolve(), not raw
    strings (see the module docstring). resolve() on a nonexistent
    path is safe on Windows/POSIX (does not require the file to
    exist)."""
    current_db = db_path().resolve()
    default_db = _default_db_path().resolve()
    out = output_path.resolve()
    default_out = _default_xlsx_path().resolve()
    if current_db != default_db and out == default_out:
        return False
    return True


def _reconfigure_stdout_utf8():
    """cp1251/other-narrow-console safety (command hygiene) -- see
    tools/hygiene_gate.py._reconfigure_stdout_utf8; this CLI prints
    filesystem paths (the xlsx output path) whose characters a narrow
    console codepage may not be able to encode."""
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass


def main(argv=None) -> int:
    _reconfigure_stdout_utf8()
    parser = argparse.ArgumentParser(
        description="Monthly token-usage statistics -> xlsx (append-only)."
    )
    parser.add_argument(
        "--output",
        default=None,
        help=(
            "Output xlsx path (default: logs/token_usage.xlsx). "
            "The ONLY mechanism for overriding the path -- see the module "
            "docstring 'OUTPUT PATH OVERRIDE + FAIL-CLOSED GUARD'."
        ),
    )
    args = parser.parse_args(argv)
    output_path = Path(args.output) if args.output else _default_xlsx_path()

    if not _is_safe_to_write(output_path):
        print(
            "REFUSED: GATEWAY_DB_PATH points to something OTHER than the "
            f"repo's default database ({_default_db_path()}), while the "
            f"output path is the default live file ({_default_xlsx_path()}) "
            "-- guard against a test/swapped context writing into the live "
            "xlsx. Pass --output <path> explicitly if this is a deliberate "
            "alternative run.",
            file=sys.stderr,
        )
        return 1

    imported, sessions, warnings = import_transcripts(transcript_glob(), db_path())
    print(f"imported {imported} new rows from {len(sessions)} sessions")
    conn = _connect(db_path())
    try:
        backfilled = backfill_costs(conn)
        conn.commit()
        if backfilled:
            print(f"backfilled costs for {backfilled} rows")
        by_family, by_project, period_sessions = aggregate(conn)
    finally:
        conn.close()
    for w in warnings:
        print(w)

    wb, ws_u, ws_p = open_or_create_workbook(output_path)
    now = datetime.now().astimezone()
    appended = []
    for wkey in sorted(by_family):
        if window_end_dt(*wkey) > now:
            continue  # window not finished yet
        label = window_label(*wkey)
        n_sessions = len(period_sessions[wkey])
        if label not in existing_periods(ws_u):
            _append_rows(ws_u, usage_rows(label, by_family[wkey], n_sessions),
                         USAGE_INT_COLS)
            appended.append(label)
        if label not in existing_periods(ws_p):
            _append_rows(ws_p, project_rows(label, by_project[wkey], n_sessions),
                         PROJECTS_INT_COLS)
    if appended:
        wb.save(output_path)
        print(f"appended periods: {', '.join(appended)} -> {output_path}")
    else:
        print("no new completed periods to append")
    return 0


if __name__ == "__main__":
    sys.exit(main())
